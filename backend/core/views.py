from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.db.models import Sum
from django.contrib.auth.models import User
from datetime import date
from decimal import Decimal
from .models import Farm, Flock, DailyEntry, Sale, FeedRate, Medication, UserProfile
from .serializers import (
    FarmSerializer, FlockSerializer, FlockListSerializer,
    DailyEntrySerializer, SaleSerializer, FeedRateSerializer,
    MedicationSerializer
)
from .permissions import IsAdmin, IsAdminOrReadOnly, CanEditFarm, CanEditFlock, CanEditFlockData


# ─── AUTH VIEWS ───

@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    from django.contrib.auth import authenticate
    username = request.data.get('username', '')
    password = request.data.get('password', '')
    user = authenticate(username=username, password=password)
    if user is None:
        return Response({'error': 'Invalid credentials'}, status=400)
    if not hasattr(user, 'profile'):
        UserProfile.objects.create(user=user, role='admin')
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(user)
    return Response({
        'access': str(refresh.access_token),
        'refresh': str(refresh),
        'user': {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'role': user.profile.role,
            'assigned_farm_ids': list(user.profile.assigned_farms.values_list('id', flat=True)),
        }
    })


@api_view(['GET'])
def me_view(request):
    user = request.user
    if not hasattr(user, 'profile'):
        UserProfile.objects.create(user=user, role='admin')
    return Response({
        'id': user.id,
        'username': user.username,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'role': user.profile.role,
        'assigned_farm_ids': list(user.profile.assigned_farms.values_list('id', flat=True)),
    })


@api_view(['POST'])
def change_password(request):
    user = request.user
    current = request.data.get('current_password', '')
    new_pass = request.data.get('new_password', '')
    if not user.check_password(current):
        return Response({'error': 'Current password is incorrect'}, status=400)
    if len(new_pass) < 6:
        return Response({'error': 'New password must be at least 6 characters'}, status=400)
    user.set_password(new_pass)
    user.save()
    # Return new tokens since password changed
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(user)
    return Response({
        'message': 'Password changed successfully',
        'access': str(refresh.access_token),
        'refresh': str(refresh),
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAdmin])
def manage_users(request):
    if request.method == 'GET':
        users = User.objects.select_related('profile').all()
        data = []
        for u in users:
            profile = getattr(u, 'profile', None)
            data.append({
                'id': u.id,
                'username': u.username,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'is_active': u.is_active,
                'role': profile.role if profile else 'admin',
                'phone': profile.phone if profile else '',
                'assigned_farm_ids': list(profile.assigned_farms.values_list('id', flat=True)) if profile else [],
            })
        return Response(data)

    # POST — create user
    username = request.data.get('username')
    password = request.data.get('password')
    first_name = request.data.get('first_name', '')
    last_name = request.data.get('last_name', '')
    role = request.data.get('role', 'supervisor')
    phone = request.data.get('phone', '')
    assigned_farm_ids = request.data.get('assigned_farm_ids', [])

    if User.objects.filter(username=username).exists():
        return Response({'error': 'Username already exists'}, status=400)

    user = User.objects.create_user(username=username, password=password,
                                     first_name=first_name, last_name=last_name)
    profile = UserProfile.objects.create(user=user, role=role, phone=phone)
    if assigned_farm_ids:
        profile.assigned_farms.set(assigned_farm_ids)
    return Response({'id': user.id, 'username': user.username}, status=201)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAdmin])
def manage_user_detail(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=404)

    if request.method == 'DELETE':
        user.delete()
        return Response(status=204)

    # PUT — update
    user.first_name = request.data.get('first_name', user.first_name)
    user.last_name = request.data.get('last_name', user.last_name)
    if request.data.get('password'):
        user.set_password(request.data['password'])
    user.is_active = request.data.get('is_active', user.is_active)
    user.save()

    if not hasattr(user, 'profile'):
        UserProfile.objects.create(user=user)
    user.profile.role = request.data.get('role', user.profile.role)
    user.profile.phone = request.data.get('phone', user.profile.phone)
    user.profile.save()
    if 'assigned_farm_ids' in request.data:
        user.profile.assigned_farms.set(request.data['assigned_farm_ids'])
    return Response({'id': user.id, 'username': user.username})


# ─── DATA VIEWSETS ───

class FarmViewSet(viewsets.ModelViewSet):
    queryset = Farm.objects.all()
    serializer_class = FarmSerializer
    permission_classes = [IsAuthenticated, CanEditFarm]

    def perform_create(self, serializer):
        # Supervisors can only create farms (admin always can)
        if not self.request.user.profile.is_admin:
            # Supervisor can create, farm gets auto-assigned to them
            farm = serializer.save()
            self.request.user.profile.assigned_farms.add(farm)
        else:
            serializer.save()


class FlockViewSet(viewsets.ModelViewSet):
    queryset = Flock.objects.select_related('farm').all()
    permission_classes = [IsAuthenticated, CanEditFlock]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return FlockSerializer
        return FlockListSerializer

    def perform_create(self, serializer):
        # Check farm permission
        farm_id = self.request.data.get('farm')
        if farm_id:
            farm = Farm.objects.get(id=farm_id)
            if not self.request.user.profile.can_edit_farm(farm):
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('You do not have permission to add flocks to this farm.')
        instance = serializer.save()
        try:
            instance.full_clean()
        except Exception as e:
            instance.delete()
            from rest_framework.exceptions import ValidationError
            raise ValidationError(e.message_dict if hasattr(e, 'message_dict') else str(e))


class DailyEntryViewSet(viewsets.ModelViewSet):
    queryset = DailyEntry.objects.select_related('flock', 'flock__farm').all()
    serializer_class = DailyEntrySerializer
    permission_classes = [IsAuthenticated, CanEditFlockData]

    def get_queryset(self):
        qs = super().get_queryset()
        flock_id = self.request.query_params.get('flock')
        if flock_id:
            qs = qs.filter(flock_id=flock_id)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        try:
            instance.full_clean()
        except Exception as e:
            instance.delete()
            from rest_framework.exceptions import ValidationError
            raise ValidationError(e.message_dict if hasattr(e, 'message_dict') else str(e))

    def perform_update(self, serializer):
        instance = serializer.save()
        try:
            instance.full_clean()
        except Exception as e:
            from rest_framework.exceptions import ValidationError
            raise ValidationError(e.message_dict if hasattr(e, 'message_dict') else str(e))


class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.select_related('flock', 'flock__farm').all()
    serializer_class = SaleSerializer
    permission_classes = [IsAuthenticated, CanEditFlockData]

    def get_queryset(self):
        qs = super().get_queryset()
        flock_id = self.request.query_params.get('flock')
        if flock_id:
            qs = qs.filter(flock_id=flock_id)
        return qs


class FeedRateViewSet(viewsets.ModelViewSet):
    queryset = FeedRate.objects.all()
    serializer_class = FeedRateSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]


class MedicationViewSet(viewsets.ModelViewSet):
    queryset = Medication.objects.all()
    serializer_class = MedicationSerializer
    permission_classes = [IsAuthenticated, CanEditFlockData]

    def get_queryset(self):
        qs = super().get_queryset()
        flock_id = self.request.query_params.get('flock')
        if flock_id:
            qs = qs.filter(flock_id=flock_id)
        return qs


def get_latest_feed_rates():
    """Get the most recent rate for each feed type."""
    rates = {}
    for ft in ['BPSC', 'BSC', 'BFP']:
        latest = FeedRate.objects.filter(feed_type=ft).first()
        if latest:
            rates[ft] = float(latest.rate_per_kg)
    return rates


@api_view(['GET'])
def dashboard(request):
    farms = Farm.objects.all()
    all_flocks = Flock.objects.all()
    active_flocks = all_flocks.filter(status='active')
    closed_flocks = all_flocks.filter(status='closed')

    total_birds_placed = all_flocks.aggregate(t=Sum('chick_count'))['t'] or 0
    total_mortality = DailyEntry.objects.aggregate(t=Sum('mortality_count'))['t'] or 0

    # Sold
    total_sold_birds = Sale.objects.aggregate(t=Sum('bird_count'))['t'] or 0
    total_sold_weight_kg = Sale.objects.aggregate(t=Sum('total_weight_kg'))['t'] or Decimal('0')
    total_sale_amount = Decimal('0')
    for sale in Sale.objects.filter(rate_per_kg__isnull=False):
        total_sale_amount += sale.total_weight_kg * sale.rate_per_kg

    # Feed by type — ALL flocks (for total feed consumed)
    BAG_KG = 50
    feed_agg = DailyEntry.objects.aggregate(
        bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
    )
    total_bpsc_bags = float(feed_agg['bpsc'] or 0)
    total_bsc_bags = float(feed_agg['bsc'] or 0)
    total_bfp_bags = float(feed_agg['bfp'] or 0)
    total_bpsc_kg = total_bpsc_bags * BAG_KG
    total_bsc_kg = total_bsc_bags * BAG_KG
    total_bfp_kg = total_bfp_bags * BAG_KG
    total_feed_kg = total_bpsc_kg + total_bsc_kg + total_bfp_kg
    total_feed_bags = total_bpsc_bags + total_bsc_bags + total_bfp_bags

    # FCR — CLOSED FLOCKS ONLY
    closed_feed_agg = DailyEntry.objects.filter(flock__status='closed').aggregate(
        bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
    )
    closed_feed_kg = (
        float(closed_feed_agg['bpsc'] or 0) +
        float(closed_feed_agg['bsc'] or 0) +
        float(closed_feed_agg['bfp'] or 0)
    ) * BAG_KG
    closed_sold_weight = float(Sale.objects.filter(flock__status='closed').aggregate(t=Sum('total_weight_kg'))['t'] or 0)

    fcr = None
    if closed_sold_weight > 0:
        fcr = round(closed_feed_kg / closed_sold_weight, 3)

    # Cost per kg — CLOSED FLOCKS ONLY
    latest_rates = get_latest_feed_rates()
    closed_bpsc_kg = float(closed_feed_agg['bpsc'] or 0) * BAG_KG
    closed_bsc_kg = float(closed_feed_agg['bsc'] or 0) * BAG_KG
    closed_bfp_kg = float(closed_feed_agg['bfp'] or 0) * BAG_KG
    closed_feed_cost = (
        closed_bpsc_kg * latest_rates.get('BPSC', 0) +
        closed_bsc_kg * latest_rates.get('BSC', 0) +
        closed_bfp_kg * latest_rates.get('BFP', 0)
    )
    cost_per_kg = None
    if closed_sold_weight > 0:
        cost_per_kg = round(closed_feed_cost / closed_sold_weight, 2)

    # Total feed cost (all flocks)
    total_feed_cost = (
        total_bpsc_kg * latest_rates.get('BPSC', 0) +
        total_bsc_kg * latest_rates.get('BSC', 0) +
        total_bfp_kg * latest_rates.get('BFP', 0)
    )

    # Today
    today = date.today()
    today_entries = DailyEntry.objects.filter(date=today)
    mortality_today = today_entries.aggregate(t=Sum('mortality_count'))['t'] or 0
    today_feed = today_entries.aggregate(
        bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
    )

    total_live_birds = sum(f.live_birds for f in active_flocks)

    data = {
        'total_birds_placed': total_birds_placed,
        'total_sold_birds': total_sold_birds,
        'total_sold_weight_kg': float(total_sold_weight_kg),
        'total_sale_amount': float(total_sale_amount),
        'total_mortality': total_mortality,
        'mortality_percentage': round((total_mortality / total_birds_placed) * 100, 2) if total_birds_placed else 0,

        'total_feed_kg': round(total_feed_kg, 2),
        'total_feed_bags': round(total_feed_bags, 2),
        'feed_by_type': {
            'bpsc_bags': round(total_bpsc_bags, 2),
            'bsc_bags': round(total_bsc_bags, 2),
            'bfp_bags': round(total_bfp_bags, 2),
            'bpsc_kg': round(total_bpsc_kg, 2),
            'bsc_kg': round(total_bsc_kg, 2),
            'bfp_kg': round(total_bfp_kg, 2),
        },
        'total_feed_cost': round(total_feed_cost, 2),
        'fcr': fcr,
        'cost_per_kg_production': cost_per_kg,
        'closed_flocks_count': closed_flocks.count(),

        'latest_feed_rates': latest_rates,

        'total_farms': farms.count(),
        'total_active_flocks': active_flocks.count(),
        'total_live_birds': total_live_birds,
        'mortality_today': mortality_today,
        'feed_today': {
            'bpsc_bags': float(today_feed['bpsc'] or 0),
            'bsc_bags': float(today_feed['bsc'] or 0),
            'bfp_bags': float(today_feed['bfp'] or 0),
        },

        'feed_rates': FeedRateSerializer(FeedRate.objects.all()[:30], many=True).data,
        'farms': FarmSerializer(farms, many=True).data,
    }
    return Response(data)


@api_view(['GET'])
def flock_cumulative(request, flock_id):
    entries = DailyEntry.objects.filter(flock_id=flock_id).order_by('date')
    flock = Flock.objects.get(id=flock_id)
    sales = Sale.objects.filter(flock_id=flock_id).order_by('date')

    BAG_KG = 50
    cum_mortality = 0
    cum_bpsc_bags = cum_bsc_bags = cum_bfp_bags = 0
    result = []

    for entry in entries:
        cum_mortality += entry.mortality_count
        cum_bpsc_bags += float(entry.feed_bpsc_bags)
        cum_bsc_bags += float(entry.feed_bsc_bags)
        cum_bfp_bags += float(entry.feed_bfp_bags)
        cum_total_bags = cum_bpsc_bags + cum_bsc_bags + cum_bfp_bags
        day_number = (entry.date - flock.placement_date).days

        result.append({
            'id': entry.id,
            'date': entry.date,
            'day_number': day_number,
            'daily_mortality': entry.mortality_count,
            'cumulative_mortality': cum_mortality,
            'mortality_percentage': round((cum_mortality / flock.chick_count) * 100, 2) if flock.chick_count else 0,
            'feed_bpsc_bags': float(entry.feed_bpsc_bags),
            'feed_bsc_bags': float(entry.feed_bsc_bags),
            'feed_bfp_bags': float(entry.feed_bfp_bags),
            'daily_feed_bags': entry.total_feed_bags,
            'daily_feed_kg': entry.total_feed_kg,
            'cumulative_feed_bags': round(cum_total_bags, 2),
            'cumulative_feed_kg': round(cum_total_bags * BAG_KG, 2),
            'cum_bpsc_bags': round(cum_bpsc_bags, 2),
            'cum_bsc_bags': round(cum_bsc_bags, 2),
            'cum_bfp_bags': round(cum_bfp_bags, 2),
            'cum_bpsc_kg': round(cum_bpsc_bags * BAG_KG, 2),
            'cum_bsc_kg': round(cum_bsc_bags * BAG_KG, 2),
            'cum_bfp_kg': round(cum_bfp_bags * BAG_KG, 2),
            'water_liters': float(entry.water_consumed_liters),
            'avg_body_weight_grams': float(entry.avg_body_weight_grams) if entry.avg_body_weight_grams else None,
        })

    return Response({
        'flock_id': flock.id,
        'farm_name': flock.farm.name,
        'placement_date': flock.placement_date,
        'chick_count': flock.chick_count,
        'age_days': flock.age_days,
        'total_mortality': flock.total_mortality,
        'total_feed_kg': float(flock.total_feed_kg),
        'feed_by_type': flock.feed_by_type,
        'feed_schedule_status': flock.feed_schedule_status,
        'total_sold_birds': flock.total_sold_birds,
        'total_sold_weight_kg': float(flock.total_sold_weight_kg),
        'live_birds': flock.live_birds,
        'fcr': flock.fcr,
        'entries': result,
        'sales': SaleSerializer(sales, many=True).data,
    })


@api_view(['GET'])
def monthly_report(request):
    """
    Monthly report: ?year=2026&month=6
    Only includes flocks that had at least one sale in that month.
    For each such flock, shows full lifecycle data: weekly mortality, total FCR,
    feed cost per kg production.
    """
    import calendar
    from datetime import timedelta

    year = int(request.query_params.get('year', date.today().year))
    month = int(request.query_params.get('month', date.today().month))

    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])

    # Find flocks that had sales this month
    flock_ids = Sale.objects.filter(
        date__gte=month_start, date__lte=month_end
    ).values_list('flock_id', flat=True).distinct()

    flocks_data = []
    BAG_KG = 50

    # Totals across all reported flocks
    grand_birds_placed = 0
    grand_mortality = 0
    grand_sold_birds = 0
    grand_sold_weight = 0
    grand_sale_amount = 0
    grand_feed_kg = 0
    grand_feed_cost = 0

    for flock in Flock.objects.filter(id__in=flock_ids).select_related('farm'):
        entries = DailyEntry.objects.filter(flock=flock).order_by('date')
        sales = Sale.objects.filter(flock=flock)
        month_sales = sales.filter(date__gte=month_start, date__lte=month_end)

        # Weekly mortality breakdown (week 1 = day 1-7, week 2 = day 8-14, etc.)
        weekly_mortality = {}
        total_mortality = 0
        for entry in entries:
            day_num = (entry.date - flock.placement_date).days
            week_num = (day_num // 7) + 1
            if week_num not in weekly_mortality:
                weekly_mortality[week_num] = {'week': week_num, 'mortality': 0, 'days': f"Day {(week_num-1)*7+1}-{week_num*7}"}
            weekly_mortality[week_num]['mortality'] += entry.mortality_count
            total_mortality += entry.mortality_count

        # Add mortality % per week
        for w in weekly_mortality.values():
            w['mortality_pct'] = round((w['mortality'] / flock.chick_count) * 100, 2) if flock.chick_count else 0

        # Feed totals
        feed_agg = entries.aggregate(
            bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
        )
        bpsc_kg = float(feed_agg['bpsc'] or 0) * BAG_KG
        bsc_kg = float(feed_agg['bsc'] or 0) * BAG_KG
        bfp_kg = float(feed_agg['bfp'] or 0) * BAG_KG
        total_feed_kg = bpsc_kg + bsc_kg + bfp_kg

        # Sale totals (full lifecycle, not just this month)
        sold_birds = sales.aggregate(t=Sum('bird_count'))['t'] or 0
        sold_weight = float(sales.aggregate(t=Sum('total_weight_kg'))['t'] or 0)
        sale_amount = 0
        for s in sales.filter(rate_per_kg__isnull=False):
            sale_amount += float(s.total_weight_kg) * float(s.rate_per_kg)

        # This month's sales
        month_sold_birds = month_sales.aggregate(t=Sum('bird_count'))['t'] or 0
        month_sold_weight = float(month_sales.aggregate(t=Sum('total_weight_kg'))['t'] or 0)
        month_sale_amount = 0
        for s in month_sales.filter(rate_per_kg__isnull=False):
            month_sale_amount += float(s.total_weight_kg) * float(s.rate_per_kg)

        # FCR
        fcr = round(total_feed_kg / sold_weight, 3) if sold_weight > 0 else None

        # Feed cost per kg production
        latest_rates = get_latest_feed_rates()
        feed_cost = (
            bpsc_kg * latest_rates.get('BPSC', 0) +
            bsc_kg * latest_rates.get('BSC', 0) +
            bfp_kg * latest_rates.get('BFP', 0)
        )
        cost_per_kg = round(feed_cost / sold_weight, 2) if sold_weight > 0 else None

        # Avg bird weight
        avg_bird_wt = round(sold_weight / sold_birds, 3) if sold_birds > 0 else None

        flock_data = {
            'flock_id': flock.id,
            'farm_name': flock.farm.name,
            'farm_code': flock.farm.farm_code,
            'farm_location': flock.farm.location,
            'placement_date': flock.placement_date,
            'chick_count': flock.chick_count,
            'age_days': flock.age_days,

            'total_mortality': total_mortality,
            'mortality_pct': round((total_mortality / flock.chick_count) * 100, 2) if flock.chick_count else 0,
            'weekly_mortality': sorted(weekly_mortality.values(), key=lambda x: x['week']),

            'total_feed_kg': round(total_feed_kg, 2),
            'feed_bpsc_kg': round(bpsc_kg, 2),
            'feed_bsc_kg': round(bsc_kg, 2),
            'feed_bfp_kg': round(bfp_kg, 2),
            'total_feed_bags': round((bpsc_kg + bsc_kg + bfp_kg) / BAG_KG, 2),

            'total_sold_birds': sold_birds,
            'total_sold_weight_kg': round(sold_weight, 2),
            'total_sale_amount': round(sale_amount, 2),
            'avg_bird_weight_kg': avg_bird_wt,

            'month_sold_birds': month_sold_birds,
            'month_sold_weight_kg': round(month_sold_weight, 2),
            'month_sale_amount': round(month_sale_amount, 2),

            'fcr': fcr,
            'feed_cost': round(feed_cost, 2),
            'cost_per_kg_production': cost_per_kg,
        }
        flocks_data.append(flock_data)

        grand_birds_placed += flock.chick_count
        grand_mortality += total_mortality
        grand_sold_birds += sold_birds
        grand_sold_weight += sold_weight
        grand_sale_amount += sale_amount
        grand_feed_kg += total_feed_kg
        grand_feed_cost += feed_cost

    grand_fcr = round(grand_feed_kg / grand_sold_weight, 3) if grand_sold_weight > 0 else None
    grand_cost_per_kg = round(grand_feed_cost / grand_sold_weight, 2) if grand_sold_weight > 0 else None

    return Response({
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'flocks_count': len(flocks_data),
        'summary': {
            'total_birds_placed': grand_birds_placed,
            'total_mortality': grand_mortality,
            'mortality_pct': round((grand_mortality / grand_birds_placed) * 100, 2) if grand_birds_placed else 0,
            'total_sold_birds': grand_sold_birds,
            'total_sold_weight_kg': round(grand_sold_weight, 2),
            'total_sale_amount': round(grand_sale_amount, 2),
            'total_feed_kg': round(grand_feed_kg, 2),
            'total_feed_cost': round(grand_feed_cost, 2),
            'fcr': grand_fcr,
            'cost_per_kg_production': grand_cost_per_kg,
        },
        'flocks': flocks_data,
    })


@api_view(['GET'])
def farm_cumulative(request, farm_id):
    """Cumulative metrics across all closed flocks for a farm."""
    farm = Farm.objects.get(id=farm_id)
    closed_flocks = Flock.objects.filter(farm=farm, status='closed')

    BAG_KG = 50
    total_birds_placed = 0
    total_mortality = 0
    total_sold_birds = 0
    total_sold_weight = 0
    total_sale_amount = 0
    total_feed_kg = 0
    total_bpsc_kg = 0
    total_bsc_kg = 0
    total_bfp_kg = 0
    total_feed_cost = 0
    flock_count = 0

    latest_rates = get_latest_feed_rates()

    for flock in closed_flocks:
        flock_count += 1
        total_birds_placed += flock.chick_count
        total_mortality += flock.total_mortality
        total_sold_birds += flock.total_sold_birds
        sold_wt = float(flock.total_sold_weight_kg)
        total_sold_weight += sold_wt

        for s in flock.sales.filter(rate_per_kg__isnull=False):
            total_sale_amount += float(s.total_weight_kg) * float(s.rate_per_kg)

        fb = flock.feed_by_type
        total_bpsc_kg += fb['bpsc_kg']
        total_bsc_kg += fb['bsc_kg']
        total_bfp_kg += fb['bfp_kg']
        total_feed_kg += fb['bpsc_kg'] + fb['bsc_kg'] + fb['bfp_kg']

        total_feed_cost += (
            fb['bpsc_kg'] * latest_rates.get('BPSC', 0) +
            fb['bsc_kg'] * latest_rates.get('BSC', 0) +
            fb['bfp_kg'] * latest_rates.get('BFP', 0)
        )

    fcr = round(total_feed_kg / total_sold_weight, 3) if total_sold_weight > 0 else None
    cost_per_kg = round(total_feed_cost / total_sold_weight, 2) if total_sold_weight > 0 else None
    avg_bird_wt = round(total_sold_weight / total_sold_birds, 3) if total_sold_birds > 0 else None

    return Response({
        'farm_id': farm.id,
        'farm_name': farm.name,
        'closed_flock_count': flock_count,
        'total_birds_placed': total_birds_placed,
        'total_mortality': total_mortality,
        'mortality_pct': round((total_mortality / total_birds_placed) * 100, 2) if total_birds_placed else 0,
        'total_sold_birds': total_sold_birds,
        'total_sold_weight_kg': round(total_sold_weight, 2),
        'avg_bird_weight_kg': avg_bird_wt,
        'total_sale_amount': round(total_sale_amount, 2),
        'total_feed_kg': round(total_feed_kg, 2),
        'total_feed_bags': round(total_feed_kg / BAG_KG, 2),
        'feed_bpsc_kg': round(total_bpsc_kg, 2),
        'feed_bsc_kg': round(total_bsc_kg, 2),
        'feed_bfp_kg': round(total_bfp_kg, 2),
        'fcr': fcr,
        'total_feed_cost': round(total_feed_cost, 2),
        'cost_per_kg_production': cost_per_kg,
    })


@api_view(['GET'])
def region_performance(request):
    """
    Performance metrics grouped by region.
    Optional: ?region=XYZ to filter one region.
    """
    region_filter = request.query_params.get('region')
    BAG_KG = 50

    farms_qs = Farm.objects.all()
    if region_filter:
        farms_qs = farms_qs.filter(region=region_filter)

    # Get all distinct regions
    all_regions = list(Farm.objects.exclude(region='').values_list('region', flat=True).distinct().order_by('region'))

    # Build per-region data
    regions_data = []
    latest_rates = get_latest_feed_rates()

    region_list = [region_filter] if region_filter else all_regions
    if not region_filter and farms_qs.filter(region='').exists():
        region_list.append('')  # include unassigned

    for region in region_list:
        r_farms = farms_qs.filter(region=region)
        r_flocks = Flock.objects.filter(farm__in=r_farms)
        r_closed = r_flocks.filter(status='closed')
        r_active = r_flocks.filter(status='active')

        birds_placed = r_flocks.aggregate(t=Sum('chick_count'))['t'] or 0
        mortality = DailyEntry.objects.filter(flock__in=r_flocks).aggregate(t=Sum('mortality_count'))['t'] or 0

        sold_birds = Sale.objects.filter(flock__in=r_flocks).aggregate(t=Sum('bird_count'))['t'] or 0
        sold_weight = float(Sale.objects.filter(flock__in=r_flocks).aggregate(t=Sum('total_weight_kg'))['t'] or 0)

        sale_amount = 0
        for s in Sale.objects.filter(flock__in=r_flocks, rate_per_kg__isnull=False):
            sale_amount += float(s.total_weight_kg) * float(s.rate_per_kg)

        # Total feed (all flocks)
        feed_agg = DailyEntry.objects.filter(flock__in=r_flocks).aggregate(
            bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
        )
        bpsc_kg = float(feed_agg['bpsc'] or 0) * BAG_KG
        bsc_kg = float(feed_agg['bsc'] or 0) * BAG_KG
        bfp_kg = float(feed_agg['bfp'] or 0) * BAG_KG
        total_feed_kg = bpsc_kg + bsc_kg + bfp_kg

        # FCR + cost/kg — CLOSED FLOCKS ONLY
        closed_feed_agg = DailyEntry.objects.filter(flock__in=r_closed).aggregate(
            bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'),
        )
        c_feed_kg = (float(closed_feed_agg['bpsc'] or 0) + float(closed_feed_agg['bsc'] or 0) + float(closed_feed_agg['bfp'] or 0)) * BAG_KG
        c_sold_weight = float(Sale.objects.filter(flock__in=r_closed).aggregate(t=Sum('total_weight_kg'))['t'] or 0)

        fcr = round(c_feed_kg / c_sold_weight, 3) if c_sold_weight > 0 else None

        c_bpsc_kg = float(closed_feed_agg['bpsc'] or 0) * BAG_KG
        c_bsc_kg = float(closed_feed_agg['bsc'] or 0) * BAG_KG
        c_bfp_kg = float(closed_feed_agg['bfp'] or 0) * BAG_KG
        feed_cost = (
            c_bpsc_kg * latest_rates.get('BPSC', 0) +
            c_bsc_kg * latest_rates.get('BSC', 0) +
            c_bfp_kg * latest_rates.get('BFP', 0)
        )
        cost_per_kg = round(feed_cost / c_sold_weight, 2) if c_sold_weight > 0 else None
        avg_bird_wt = round(sold_weight / sold_birds, 3) if sold_birds > 0 else None

        live_birds = sum(f.live_birds for f in r_active)

        regions_data.append({
            'region': region or 'Unassigned',
            'farm_count': r_farms.count(),
            'active_flocks': r_active.count(),
            'closed_flocks': r_closed.count(),
            'live_birds': live_birds,
            'total_birds_placed': birds_placed,
            'total_mortality': mortality,
            'mortality_pct': round((mortality / birds_placed) * 100, 2) if birds_placed else 0,
            'total_sold_birds': sold_birds,
            'total_sold_weight_kg': round(sold_weight, 2),
            'avg_bird_weight_kg': avg_bird_wt,
            'total_sale_amount': round(sale_amount, 2),
            'total_feed_kg': round(total_feed_kg, 2),
            'total_feed_bags': round(total_feed_kg / BAG_KG, 2),
            'fcr': fcr,
            'total_feed_cost': round(feed_cost, 2),
            'cost_per_kg_production': cost_per_kg,
        })

    return Response({
        'regions': all_regions,
        'data': regions_data,
    })


@api_view(['GET'])
def export_monthly_report(request):
    """Export monthly report as Excel. ?year=2026&month=6"""
    import calendar
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    year = int(request.query_params.get('year', date.today().year))
    month = int(request.query_params.get('month', date.today().month))
    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])

    flock_ids = Sale.objects.filter(
        date__gte=month_start, date__lte=month_end
    ).values_list('flock_id', flat=True).distinct()

    BAG_KG = 50
    latest_rates = get_latest_feed_rates()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:O1')
    ws['A1'] = f"Monthly Report — {calendar.month_name[month]} {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Farm', 'Farm Code', 'Placed', 'Chicks', 'Age', 'Mortality', 'Mort%',
               'Sold (birds)', 'Sold (kg)', 'Avg Wt', 'Feed (kg)', 'Feed (bags)',
               'FCR', 'Feed Cost (₹)', 'Cost/kg (₹)', 'Sale Amt (₹)', 'Med Cost (₹)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row = 4
    grand = {'birds': 0, 'mort': 0, 'sold_b': 0, 'sold_w': 0, 'feed': 0, 'feed_cost': 0, 'sale': 0, 'med': 0}

    for flock in Flock.objects.filter(id__in=flock_ids).select_related('farm'):
        entries = DailyEntry.objects.filter(flock=flock)
        sales = Sale.objects.filter(flock=flock)

        mortality = entries.aggregate(t=Sum('mortality_count'))['t'] or 0
        feed_agg = entries.aggregate(bpsc=Sum('feed_bpsc_bags'), bsc=Sum('feed_bsc_bags'), bfp=Sum('feed_bfp_bags'))
        feed_kg = (float(feed_agg['bpsc'] or 0) + float(feed_agg['bsc'] or 0) + float(feed_agg['bfp'] or 0)) * BAG_KG
        bpsc_kg = float(feed_agg['bpsc'] or 0) * BAG_KG
        bsc_kg = float(feed_agg['bsc'] or 0) * BAG_KG
        bfp_kg = float(feed_agg['bfp'] or 0) * BAG_KG

        sold_birds = sales.aggregate(t=Sum('bird_count'))['t'] or 0
        sold_weight = float(sales.aggregate(t=Sum('total_weight_kg'))['t'] or 0)
        sale_amount = sum(float(s.total_weight_kg) * float(s.rate_per_kg) for s in sales.filter(rate_per_kg__isnull=False))
        med_cost = float(flock.total_medication_cost)

        fcr = round(feed_kg / sold_weight, 3) if sold_weight > 0 else None
        feed_cost = bpsc_kg * latest_rates.get('BPSC', 0) + bsc_kg * latest_rates.get('BSC', 0) + bfp_kg * latest_rates.get('BFP', 0)
        cost_per_kg = round(feed_cost / sold_weight, 2) if sold_weight > 0 else None
        avg_wt = round(sold_weight / sold_birds, 3) if sold_birds > 0 else None

        values = [
            flock.farm.name, flock.farm.farm_code, str(flock.placement_date), flock.chick_count,
            flock.age_days, mortality, f"{round((mortality/flock.chick_count)*100,2) if flock.chick_count else 0}%",
            sold_birds, round(sold_weight, 2), avg_wt or '', round(feed_kg, 2), round(feed_kg / BAG_KG, 1),
            fcr or '', round(feed_cost, 2), cost_per_kg or '', round(sale_amount, 2), round(med_cost, 2)
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
        row += 1

        grand['birds'] += flock.chick_count
        grand['mort'] += mortality
        grand['sold_b'] += sold_birds
        grand['sold_w'] += sold_weight
        grand['feed'] += feed_kg
        grand['feed_cost'] += feed_cost
        grand['sale'] += sale_amount
        grand['med'] += med_cost

    # Totals row
    g_fcr = round(grand['feed'] / grand['sold_w'], 3) if grand['sold_w'] > 0 else ''
    g_cpk = round(grand['feed_cost'] / grand['sold_w'], 2) if grand['sold_w'] > 0 else ''
    g_avg = round(grand['sold_w'] / grand['sold_b'], 3) if grand['sold_b'] > 0 else ''
    totals = [
        'TOTAL', '', '', grand['birds'], '',
        grand['mort'], f"{round((grand['mort']/grand['birds'])*100,2) if grand['birds'] else 0}%",
        grand['sold_b'], round(grand['sold_w'], 2), g_avg, round(grand['feed'], 2),
        round(grand['feed'] / BAG_KG, 1), g_fcr, round(grand['feed_cost'], 2), g_cpk,
        round(grand['sale'], 2), round(grand['med'], 2)
    ]
    for col, v in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 20)

    # Write response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="monthly_report_{year}_{month:02d}.xlsx"'
    return response


@api_view(['GET'])
def export_flock_report(request, flock_id):
    """Export single flock data as Excel."""
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    flock = Flock.objects.select_related('farm').get(id=flock_id)
    entries = DailyEntry.objects.filter(flock=flock).order_by('date')
    sales = Sale.objects.filter(flock=flock).order_by('date')
    medications = flock.medications.order_by('date')

    BAG_KG = 50
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    info = [
        ('Farm', flock.farm.name), ('Farm Code', flock.farm.farm_code),
        ('Placement Date', str(flock.placement_date)), ('Chicks Placed', flock.chick_count),
        ('Age (days)', flock.age_days), ('Status', flock.status),
        ('Live Birds', flock.live_birds), ('Total Mortality', flock.total_mortality),
        ('Mortality %', f"{flock.mortality_percentage}%"),
        ('Total Feed (kg)', float(flock.total_feed_kg)),
        ('Sold Birds', flock.total_sold_birds),
        ('Sold Weight (kg)', float(flock.total_sold_weight_kg)),
        ('FCR', flock.fcr or ''), ('Medication Cost (₹)', float(flock.total_medication_cost)),
    ]
    for r, (label, val) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)

    # Sheet 2: Daily Entries
    ws2 = wb.create_sheet("Daily Entries")
    headers = ['Day', 'Date', 'Mortality', 'BPSC (bags)', 'BSC (bags)', 'BFP (bags)', 'Total (bags)', 'Total (kg)', 'Water (L)', 'Body Wt (g)', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for r, entry in enumerate(entries, 2):
        day = (entry.date - flock.placement_date).days
        vals = [day, str(entry.date), entry.mortality_count,
                float(entry.feed_bpsc_bags), float(entry.feed_bsc_bags), float(entry.feed_bfp_bags),
                entry.total_feed_bags, entry.total_feed_kg,
                float(entry.water_consumed_liters),
                float(entry.avg_body_weight_grams) if entry.avg_body_weight_grams else '',
                entry.notes]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=r, column=col, value=v).border = thin_border

    # Sheet 3: Sales
    ws3 = wb.create_sheet("Sales")
    s_headers = ['Date', 'Birds', 'Weight (kg)', 'Avg/Bird (kg)', 'Rate (₹/kg)', 'Amount (₹)', 'Notes']
    for col, h in enumerate(s_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for r, sale in enumerate(sales, 2):
        vals = [str(sale.date), sale.bird_count, float(sale.total_weight_kg),
                sale.avg_bird_weight_kg, float(sale.rate_per_kg) if sale.rate_per_kg else '',
                sale.total_amount or '', sale.notes]
        for col, v in enumerate(vals, 1):
            ws3.cell(row=r, column=col, value=v).border = thin_border

    # Sheet 4: Medications
    ws4 = wb.create_sheet("Medications")
    m_headers = ['Date', 'Name', 'Dose', 'Route', 'Cost (₹)', 'Reason']
    for col, h in enumerate(m_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for r, med in enumerate(medications, 2):
        vals = [str(med.date), med.name, med.dose, med.get_route_display() if med.route else '', float(med.cost), med.reason]
        for col, v in enumerate(vals, 1):
            ws4.cell(row=r, column=col, value=v).border = thin_border

    # Auto-width all sheets
    for sheet in wb.sheetnames:
        for col in wb[sheet].columns:
            max_len = max(len(str(c.value or '')) for c in col) + 2
            wb[sheet].column_dimensions[col[0].column_letter].width = min(max_len, 25)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fname = f"flock_{flock.farm.farm_code}_{flock.placement_date}.xlsx"
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
